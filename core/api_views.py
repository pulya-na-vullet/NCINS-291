import json
import csv
from io import StringIO

from django.db.models import Q
from django.db import transaction
from django.http import HttpResponse
from django.http import JsonResponse as DjangoJsonResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods

from .models import (
    AIActivityLog,
    AIAnalysis,
    Comment,
    Project,
    ProjectIntegrationSettings,
    Section,
    Tag,
    TestCase,
    TestCaseReview,
    TestCaseUserStory,
    TestRun,
    TestRunTestCase,
    TestSuiteAIReviewJob,
    TestStep,
    TestSuite,
    TraceabilityMatrix,
    User,
    UserStory,
)
from .serializers import (
    CommentSerializer,
    ProjectSerializer,
    SectionSerializer,
    TagSerializer,
    TestCaseReviewSerializer,
    TestCaseSerializer,
    TestRunSerializer,
    TestRunTestCaseSerializer,
    TestStepSerializer,
    TestSuiteSerializer,
    TraceabilityMatrixSerializer,
    UserSerializer,
    UserStorySerializer,
)
from .services import (
    analyze_traceability_model_quality,
    build_api_response,
    create_or_update_review,
    create_test_run,
    generate_and_store_matrix,
    import_test_cases_from_provider,
    normalize_status,
    run_ai_test_suite_analysis,
    test_ai_provider_connection,
)


def _to_camel_case(value):
    if isinstance(value, list):
        return [_to_camel_case(v) for v in value]
    if isinstance(value, dict):
        transformed = {}
        for key, val in value.items():
            if isinstance(key, str):
                parts = key.split("_")
                camel_key = parts[0] + "".join(part[:1].upper() + part[1:] for part in parts[1:])
            else:
                camel_key = key
            transformed[camel_key] = _to_camel_case(val)
        return transformed
    return value


def JsonResponse(data, *args, **kwargs):
    return DjangoJsonResponse(_to_camel_case(data), *args, **kwargs)


def _json_body(request):
    if not request.body:
        return {}
    return json.loads(request.body.decode("utf-8"))


def _is_analyst_request(request):
    if not getattr(request, "user", None) or not request.user.is_authenticated:
        return False
    core_user = User.objects.filter(username=request.user.username, enabled=True).first()
    if not core_user:
        return False
    roles = set(core_user.roles or [])
    return "ANALYST" in roles or "ROLE_ANALYST" in roles


def _parse_business_criticality(payload):
    raw_value = payload.get("businessCriticality", payload.get("business_criticality"))
    if raw_value in (None, ""):
        return None, None
    try:
        value = int(raw_value)
    except (TypeError, ValueError):
        return None, "Критичность бизнеса должна быть числом от 1 до 10"
    if value < 1 or value > 10:
        return None, "Критичность бизнеса должна быть в диапазоне от 1 до 10"
    return value, None


def _resolve_comment_user_id(request):
    if not getattr(request, "user", None) or not request.user.is_authenticated:
        return None
    core_user = User.objects.filter(username=request.user.username).first()
    return core_user.id if core_user else None


def _resolve_core_user(request):
    if not getattr(request, "user", None) or not request.user.is_authenticated:
        return None
    return User.objects.filter(username=request.user.username, enabled=True).first()


def _is_admin_user(core_user):
    if not core_user:
        return False
    return "ADMIN" in (core_user.roles or [])


def _serialize_ai_review_job(job: TestSuiteAIReviewJob):
    percent = 0.0
    if job.total_cases:
        percent = round((job.processed_cases * 100.0) / job.total_cases, 2)
    return {
        "id": job.id,
        "test_suite_id": job.test_suite_id,
        "status": job.status,
        "total_cases": job.total_cases,
        "processed_cases": job.processed_cases,
        "success_cases": job.success_cases,
        "failed_cases": job.failed_cases,
        "remaining_cases": len(job.queue_case_ids or []),
        "progress_percent": percent,
        "started_by_id": job.started_by_id,
        "started_by_name": job.started_by.full_name if job.started_by else None,
        "started_at": job.started_at,
        "finished_at": job.finished_at,
        "last_error": job.last_error,
    }


def _create_ai_activity_log(
    *,
    action_type: str,
    status: str,
    initiated_by=None,
    project=None,
    test_suite=None,
    test_case=None,
    message: str | None = None,
):
    return AIActivityLog.objects.create(
        action_type=action_type,
        status=status,
        initiated_by=initiated_by,
        project=project,
        test_suite=test_suite,
        test_case=test_case,
        message=message,
        started_at=timezone.now(),
    )


@require_http_methods(["GET"])
def projects_list(request):
    return JsonResponse(ProjectSerializer(Project.objects.all(), many=True).data, safe=False)


@csrf_exempt
@require_http_methods(["POST"])
def projects_create(request):
    payload = _json_body(request)
    name = (payload.get("name") or "").strip()
    if not name:
        return JsonResponse(build_api_response(False, "Название проекта не может быть пустым!"))
    project = Project.objects.create(name=name)
    return JsonResponse(build_api_response(True, "Проект создан успешно!", project=ProjectSerializer(project).data))


@require_http_methods(["GET"])
def projects_get(request, id):
    project = Project.objects.filter(id=id).first()
    if not project:
        return JsonResponse(build_api_response(False, "Проект не найден!"))
    return JsonResponse(build_api_response(True, project=ProjectSerializer(project).data))


@csrf_exempt
@require_http_methods(["PUT"])
def projects_update(request, id):
    project = Project.objects.filter(id=id).first()
    if not project:
        return JsonResponse(build_api_response(False, "Проект не найден!"))
    payload = _json_body(request)
    name = (payload.get("name") or "").strip()
    if not name:
        return JsonResponse(build_api_response(False, "Название проекта не может быть пустым!"))
    project.name = name
    project.save()
    return JsonResponse(build_api_response(True, "Проект обновлен успешно!", project=ProjectSerializer(project).data))


@csrf_exempt
@require_http_methods(["DELETE"])
def projects_delete(request, id):
    deleted, _ = Project.objects.filter(id=id).delete()
    return JsonResponse(build_api_response(bool(deleted), "Проект удален успешно!" if deleted else "Проект не найден!"))


@require_http_methods(["GET"])
def project_test_suites(request, project_id):
    data = TestSuiteSerializer(TestSuite.objects.filter(project_id=project_id), many=True).data
    return JsonResponse(data, safe=False)


@csrf_exempt
@require_http_methods(["POST"])
def project_test_suites_create(request, project_id):
    payload = _json_body(request)
    name = (payload.get("name") or "").strip()
    if not name:
        return JsonResponse(build_api_response(False, "Название тест-сьюта не может быть пустым!"))
    if not Project.objects.filter(id=project_id).exists():
        return JsonResponse(build_api_response(False, "Проект не найден!"))
    suite = TestSuite.objects.create(name=name, project_id=project_id)
    return JsonResponse(build_api_response(True, "Тест-сьют создан успешно!", testSuite=TestSuiteSerializer(suite).data))


@require_http_methods(["GET"])
def test_suite_get(request, id):
    suite = TestSuite.objects.filter(id=id).first()
    if not suite:
        return JsonResponse(build_api_response(False, "Тест-сьют не найден!"))
    return JsonResponse(build_api_response(True, testSuite=TestSuiteSerializer(suite).data))


@csrf_exempt
@require_http_methods(["PUT"])
def test_suite_update(request, id):
    suite = TestSuite.objects.filter(id=id).first()
    if not suite:
        return JsonResponse(build_api_response(False, "Тест-сьют не найден!"))
    payload = _json_body(request)
    name = (payload.get("name") or "").strip()
    if not name:
        return JsonResponse(build_api_response(False, "Название тест-сьюта не может быть пустым!"))
    suite.name = name
    suite.save()
    return JsonResponse(build_api_response(True, "Тест-сьют обновлен успешно!", testSuite=TestSuiteSerializer(suite).data))


@csrf_exempt
@require_http_methods(["DELETE"])
def test_suite_delete(request, id):
    deleted, _ = TestSuite.objects.filter(id=id).delete()
    return JsonResponse(build_api_response(bool(deleted), "Тест-сьют удален успешно!" if deleted else "Тест-сьют не найден!"))


@require_http_methods(["GET"])
def project_test_suites_search(request, project_id):
    term = request.GET.get("term", "")
    suites = TestSuite.objects.filter(project_id=project_id, name__icontains=term)
    return JsonResponse(TestSuiteSerializer(suites, many=True).data, safe=False)


@require_http_methods(["GET"])
def test_suite_cases(request, test_suite_id):
    queryset = TestCase.objects.filter(test_suite_id=test_suite_id)
    return JsonResponse(TestCaseSerializer(queryset, many=True).data, safe=False)


@require_http_methods(["GET"])
def project_cases(request, project_id):
    queryset = TestCase.objects.filter(test_suite__project_id=project_id)
    return JsonResponse(TestCaseSerializer(queryset, many=True).data, safe=False)


@csrf_exempt
@require_http_methods(["POST"])
def test_suite_cases_create(request, test_suite_id):
    payload = _json_body(request)
    name = (payload.get("name") or "").strip()
    if not name:
        return JsonResponse(build_api_response(False, "Название тест-кейса не может быть пустым!"))
    if not TestSuite.objects.filter(id=test_suite_id).exists():
        return JsonResponse(build_api_response(False, "Тест-сьют не найден!"))
    case = TestCase.objects.create(
        test_suite_id=test_suite_id,
        name=name,
        description=payload.get("description"),
        preconditions=payload.get("preconditions"),
    )
    return JsonResponse(build_api_response(True, "Тест-кейс создан успешно!", testCase=TestCaseSerializer(case).data))


@require_http_methods(["GET"])
def test_case_get(request, id):
    case = TestCase.objects.filter(id=id).first()
    if not case:
        return JsonResponse(build_api_response(False, "Тест-кейс не найден!"))
    return JsonResponse(build_api_response(True, testCase=TestCaseSerializer(case).data))


@csrf_exempt
@require_http_methods(["PUT"])
def test_case_update(request, id):
    case = TestCase.objects.filter(id=id).first()
    if not case:
        return JsonResponse(build_api_response(False, "Тест-кейс не найден!"))
    payload = _json_body(request)
    case.name = payload.get("name") or case.name
    case.description = payload.get("description")
    case.preconditions = payload.get("preconditions")
    case.priority = normalize_status(payload.get("priority"), [v for v, _ in TestCase.Priority.choices], case.priority)
    case.status = normalize_status(payload.get("status"), [v for v, _ in TestCase.Status.choices], case.status)
    case.save()
    return JsonResponse(build_api_response(True, "Тест-кейс обновлен успешно!", testCase=TestCaseSerializer(case).data))


@csrf_exempt
@require_http_methods(["DELETE"])
def test_case_delete(request, id):
    deleted, _ = TestCase.objects.filter(id=id).delete()
    return JsonResponse(build_api_response(bool(deleted), "Тест-кейс удален успешно!" if deleted else "Тест-кейс не найден!"))


@csrf_exempt
@require_http_methods(["POST"])
def test_case_clone(request, id):
    source = TestCase.objects.filter(id=id).first()
    if not source:
        return JsonResponse(build_api_response(False, "Тест-кейс не найден!"))
    clone = TestCase.objects.create(
        test_suite_id=source.test_suite_id,
        name=f"{source.name} (copy)",
        description=source.description,
        preconditions=source.preconditions,
        priority=source.priority,
        status=source.status,
    )
    clone.tags.set(source.tags.all())
    for step in source.steps.all():
        TestStep.objects.create(
            test_case=clone,
            step_number=step.step_number,
            action=step.action,
            expected_result=step.expected_result,
        )
    for link in source.test_case_user_stories.all():
        TestCaseUserStory.objects.get_or_create(
            test_case=clone,
            user_story=link.user_story,
            defaults={"project_id": link.project_id or clone.test_suite.project_id},
        )
    return JsonResponse(build_api_response(True, "Тест-кейс склонирован успешно!", testCase=TestCaseSerializer(clone).data))


@require_http_methods(["GET"])
def test_suite_cases_search(request, test_suite_id):
    term = request.GET.get("term", "")
    queryset = TestCase.objects.filter(test_suite_id=test_suite_id).filter(
        Q(name__icontains=term) | Q(description__icontains=term) | Q(preconditions__icontains=term)
    )
    return JsonResponse(TestCaseSerializer(queryset, many=True).data, safe=False)


@csrf_exempt
@require_http_methods(["POST"])
def import_test_cases_external(request, test_suite_id):
    payload = _json_body(request)
    provider = (payload.get("provider") or "").strip().lower()
    if provider not in {"allure", "testit"}:
        return JsonResponse(build_api_response(False, "Укажите provider: allure или testit"), status=400)
    result = import_test_cases_from_provider(test_suite_id, provider)
    status = 200 if result.get("success") else 400
    return JsonResponse(result, status=status)


@csrf_exempt
@require_http_methods(["POST"])
def test_case_add_tag(request, test_case_id, tag_id):
    case = get_object_or_404(TestCase, id=test_case_id)
    tag = get_object_or_404(Tag, id=tag_id)
    case.tags.add(tag)
    return JsonResponse(build_api_response(True, "Тег добавлен к тест-кейсу!", testCase=TestCaseSerializer(case).data))


@csrf_exempt
@require_http_methods(["DELETE"])
def test_case_remove_tag(request, test_case_id, tag_id):
    case = get_object_or_404(TestCase, id=test_case_id)
    tag = get_object_or_404(Tag, id=tag_id)
    case.tags.remove(tag)
    return JsonResponse(build_api_response(True, "Тег удален из тест-кейса!", testCase=TestCaseSerializer(case).data))


@csrf_exempt
@require_http_methods(["POST"])
def test_case_add_step(request, test_case_id):
    case = get_object_or_404(TestCase, id=test_case_id)
    payload = _json_body(request)
    step = TestStep.objects.create(
        test_case=case,
        step_number=payload.get("stepNumber") or case.steps.count() + 1,
        action=payload.get("action") or "",
        expected_result=payload.get("expectedResult"),
    )
    return JsonResponse(build_api_response(True, "Шаг добавлен к тест-кейсу!", step=TestStepSerializer(step).data, testCase=TestCaseSerializer(case).data))


@csrf_exempt
@require_http_methods(["PUT"])
def test_step_update(request, step_id):
    step = get_object_or_404(TestStep, id=step_id)
    payload = _json_body(request)
    step.step_number = payload.get("stepNumber") or step.step_number
    step.action = payload.get("action") or step.action
    step.expected_result = payload.get("expectedResult")
    step.save()
    return JsonResponse(build_api_response(True, "Шаг обновлен!", testCase=TestCaseSerializer(step.test_case).data))


@csrf_exempt
@require_http_methods(["DELETE"])
def test_step_delete(request, step_id):
    step = get_object_or_404(TestStep, id=step_id)
    test_case = step.test_case
    step.delete()
    return JsonResponse(build_api_response(True, "Шаг удален!", testCase=TestCaseSerializer(test_case).data))


@csrf_exempt
@require_http_methods(["POST"])
def test_case_add_user_story(request, test_case_id, user_story_id):
    case = get_object_or_404(TestCase, id=test_case_id)
    story = get_object_or_404(UserStory, id=user_story_id)
    link, _ = TestCaseUserStory.objects.get_or_create(
        test_case=case,
        user_story=story,
        defaults={"project_id": case.test_suite.project_id},
    )
    if not link.project_id:
        link.project_id = case.test_suite.project_id
        link.save()
    return JsonResponse(build_api_response(True, "User story добавлена к тест-кейсу!", testCase=TestCaseSerializer(case).data))


@csrf_exempt
@require_http_methods(["DELETE"])
def test_case_remove_user_story(request, test_case_id, user_story_id):
    case = get_object_or_404(TestCase, id=test_case_id)
    TestCaseUserStory.objects.filter(test_case_id=test_case_id, user_story_id=user_story_id).delete()
    return JsonResponse(build_api_response(True, "User story удалена из тест-кейса!", testCase=TestCaseSerializer(case).data))


@require_http_methods(["GET"])
def project_sections(request, project_id):
    return JsonResponse(SectionSerializer(Section.objects.filter(project_id=project_id), many=True).data, safe=False)


@csrf_exempt
@require_http_methods(["POST"])
def project_sections_create(request, project_id):
    payload = _json_body(request)
    name = (payload.get("name") or "").strip()
    if not name:
        return JsonResponse(build_api_response(False, "Название секции не может быть пустым!"))
    section = Section.objects.create(name=name, project_id=project_id)
    return JsonResponse(build_api_response(True, "Секция создана успешно!", section=SectionSerializer(section).data))


@csrf_exempt
@require_http_methods(["POST"])
def project_sections_import(request, project_id):
    payload = _json_body(request)
    import_data = payload.get("importData") or []
    imported_sections = 0
    imported_user_stories = 0
    for row in import_data:
        section_name = (row.get("sectionName") or row.get("section") or "").strip()
        story_name = (row.get("userStoryName") or row.get("userStory") or "").strip()
        if not section_name:
            continue
        section, created = Section.objects.get_or_create(project_id=project_id, name=section_name)
        if created:
            imported_sections += 1
        if story_name:
            _, us_created = UserStory.objects.get_or_create(section=section, name=story_name)
            if us_created:
                imported_user_stories += 1
    return JsonResponse(
        build_api_response(
            True,
            "Импорт успешно завершен!",
            importedSections=imported_sections,
            importedUserStories=imported_user_stories,
        )
    )


@require_http_methods(["GET"])
def section_get(request, id):
    section = Section.objects.filter(id=id).first()
    if not section:
        return JsonResponse(build_api_response(False, "Секция не найдена!"))
    return JsonResponse(build_api_response(True, section=SectionSerializer(section).data))


@csrf_exempt
@require_http_methods(["PUT"])
def section_update(request, id):
    section = Section.objects.filter(id=id).first()
    if not section:
        return JsonResponse(build_api_response(False, "Секция не найдена!"))
    payload = _json_body(request)
    name = (payload.get("name") or "").strip()
    if not name:
        return JsonResponse(build_api_response(False, "Название секции не может быть пустым!"))
    section.name = name
    section.save()
    return JsonResponse(build_api_response(True, "Секция обновлена успешно!", section=SectionSerializer(section).data))


@csrf_exempt
@require_http_methods(["DELETE"])
def section_delete(request, id):
    deleted, _ = Section.objects.filter(id=id).delete()
    return JsonResponse(build_api_response(bool(deleted), "Секция удалена успешно!" if deleted else "Секция не найдена!"))


@csrf_exempt
@require_http_methods(["DELETE"])
def section_delete_with_user_stories(request, id):
    deleted, _ = Section.objects.filter(id=id).delete()
    return JsonResponse(
        build_api_response(bool(deleted), "Секция и все связанные user stories успешно удалены!" if deleted else "Секция не найдена!")
    )


@require_http_methods(["GET"])
def section_user_stories(request, section_id):
    return JsonResponse(UserStorySerializer(UserStory.objects.filter(section_id=section_id), many=True).data, safe=False)


@csrf_exempt
@require_http_methods(["POST"])
def section_user_stories_create(request, section_id):
    payload = _json_body(request)
    name = (payload.get("name") or "").strip()
    if not name:
        return JsonResponse(build_api_response(False, "Название user story не может быть пустым!"))
    criticality, error = _parse_business_criticality(payload)
    if error:
        return JsonResponse(build_api_response(False, error))
    if criticality is not None and not _is_analyst_request(request):
        return JsonResponse(build_api_response(False, "Только аналитик может задавать критичность бизнеса"), status=403)
    story = UserStory.objects.create(
        section_id=section_id,
        name=name,
        business_criticality=criticality,
    )
    return JsonResponse(build_api_response(True, "User story создана успешно!", userStory=UserStorySerializer(story).data))


@csrf_exempt
@require_http_methods(["POST"])
def section_user_stories_import(request, section_id):
    payload = _json_body(request)
    user_stories = payload.get("userStories") or []
    imported_count = 0
    for value in user_stories:
        if (value or "").strip():
            _, created = UserStory.objects.get_or_create(section_id=section_id, name=value.strip())
            imported_count += int(created)
    return JsonResponse(build_api_response(True, f"Успешно импортировано {imported_count} user stories", importedCount=imported_count))


@require_http_methods(["GET"])
def user_story_get(request, id):
    story = UserStory.objects.filter(id=id).first()
    if not story:
        return JsonResponse(build_api_response(False, "User story не найдена!"))
    return JsonResponse(build_api_response(True, userStory=UserStorySerializer(story).data))


@csrf_exempt
@require_http_methods(["PUT"])
def user_story_update(request, id):
    story = UserStory.objects.filter(id=id).first()
    if not story:
        return JsonResponse(build_api_response(False, "User story не найдена!"))
    payload = _json_body(request)
    name = (payload.get("name") or "").strip()
    if not name:
        return JsonResponse(build_api_response(False, "Название user story не может быть пустым!"))
    criticality, error = _parse_business_criticality(payload)
    if error:
        return JsonResponse(build_api_response(False, error))
    if "businessCriticality" in payload or "business_criticality" in payload:
        if not _is_analyst_request(request):
            return JsonResponse(build_api_response(False, "Только аналитик может изменять критичность бизнеса"), status=403)
        story.business_criticality = criticality
    story.name = name
    story.save()
    return JsonResponse(build_api_response(True, "User story обновлена успешно!", userStory=UserStorySerializer(story).data))


@csrf_exempt
@require_http_methods(["DELETE"])
def user_story_delete(request, id):
    deleted, _ = UserStory.objects.filter(id=id).delete()
    return JsonResponse(build_api_response(bool(deleted), "User story удалена успешно!" if deleted else "User story не найдена!"))


@require_http_methods(["GET"])
def project_user_stories(request, project_id):
    stories = UserStory.objects.filter(section__project_id=project_id)
    return JsonResponse(UserStorySerializer(stories, many=True).data, safe=False)


@require_http_methods(["GET"])
def project_tags(request, project_id):
    return JsonResponse(TagSerializer(Tag.objects.filter(project_id=project_id), many=True).data, safe=False)


@csrf_exempt
@require_http_methods(["POST"])
def project_tags_create(request, project_id):
    payload = _json_body(request)
    name = (payload.get("name") or "").strip()
    if not name:
        return JsonResponse(build_api_response(False, "Название тега не может быть пустым!"))
    tag = Tag.objects.create(project_id=project_id, name=name, color=payload.get("color") or "#6c757d")
    return JsonResponse(build_api_response(True, "Тег создан успешно!", tag=TagSerializer(tag).data))


@require_http_methods(["GET"])
def tag_get(request, id):
    tag = Tag.objects.filter(id=id).first()
    if not tag:
        return JsonResponse(build_api_response(False, "Тег не найден!"))
    return JsonResponse(build_api_response(True, tag=TagSerializer(tag).data))


@csrf_exempt
@require_http_methods(["PUT"])
def tag_update(request, id):
    tag = Tag.objects.filter(id=id).first()
    if not tag:
        return JsonResponse(build_api_response(False, "Тег не найден!"))
    payload = _json_body(request)
    name = (payload.get("name") or "").strip()
    if not name:
        return JsonResponse(build_api_response(False, "Название тега не может быть пустым!"))
    tag.name = name
    tag.color = payload.get("color") or "#6c757d"
    tag.save()
    return JsonResponse(build_api_response(True, "Тег обновлен успешно!", tag=TagSerializer(tag).data))


@csrf_exempt
@require_http_methods(["DELETE"])
def tag_delete(request, id):
    deleted, _ = Tag.objects.filter(id=id).delete()
    return JsonResponse(build_api_response(bool(deleted), "Тег удален успешно!" if deleted else "Тег не найден!"))


@csrf_exempt
@require_http_methods(["POST"])
def comments_create(request):
    payload = _json_body(request)
    current_user_id = _resolve_comment_user_id(request)
    # For authenticated users always pin author to current account.
    user_id = current_user_id if current_user_id is not None else payload.get("userId")
    comment = Comment.objects.create(
        test_case_id=payload.get("testCaseId"),
        content=payload.get("content") or "",
        user_id=user_id,
        comment_type=Comment.CommentType.MANUAL,
    )
    return JsonResponse(build_api_response(True, "Комментарий добавлен", comment=CommentSerializer(comment).data))


@csrf_exempt
@require_http_methods(["PUT"])
def comments_update(request, comment_id):
    comment = get_object_or_404(Comment, id=comment_id)
    core_user = _resolve_core_user(request)
    is_admin = _is_admin_user(core_user)
    if comment.user_id:
        if not core_user:
            return JsonResponse(build_api_response(False, "Для редактирования комментария нужна авторизация."))
        if not is_admin and comment.user_id != core_user.id:
            return JsonResponse(build_api_response(False, "Редактировать можно только собственный комментарий."))
    elif not is_admin:
        return JsonResponse(build_api_response(False, "Комментарий без автора может изменить только администратор."))

    payload = _json_body(request)
    comment.content = payload.get("content") or comment.content
    comment.save()
    return JsonResponse(build_api_response(True, "Комментарий обновлен", comment=CommentSerializer(comment).data))


@csrf_exempt
@require_http_methods(["DELETE"])
def comments_delete(request, comment_id):
    comment = get_object_or_404(Comment, id=comment_id)
    core_user = _resolve_core_user(request)
    is_admin = _is_admin_user(core_user)
    if comment.user_id:
        if not core_user:
            return JsonResponse(build_api_response(False, "Для удаления комментария нужна авторизация."))
        if not is_admin and comment.user_id != core_user.id:
            return JsonResponse(build_api_response(False, "Удалять можно только собственный комментарий."))
    elif not is_admin:
        return JsonResponse(build_api_response(False, "Комментарий без автора может удалить только администратор."))

    comment.delete()
    return JsonResponse(build_api_response(True, "Комментарий удален"))


@require_http_methods(["GET"])
def comments_by_test_case(request, test_case_id):
    comments = Comment.objects.filter(test_case_id=test_case_id)
    return JsonResponse(build_api_response(True, comments=CommentSerializer(comments, many=True).data))


@csrf_exempt
@require_http_methods(["POST"])
def comments_ai_create(request, test_case_id):
    payload = _json_body(request)
    content = (payload.get("content") or "").strip()
    if not content:
        return JsonResponse(build_api_response(False, "Содержание комментария не может быть пустым"), status=400)
    comment = Comment.objects.create(
        test_case_id=test_case_id,
        content=content,
        comment_type=Comment.CommentType.AI_GENERATED,
    )
    return JsonResponse(build_api_response(True, "AI комментарий добавлен", comment=CommentSerializer(comment).data))


@require_http_methods(["GET"])
def users_list(request):
    users = User.objects.all()
    return JsonResponse({"success": True, "message": "Пользователи получены успешно", "data": UserSerializer(users, many=True).data})


@require_http_methods(["GET"])
def users_get(request, id):
    user = User.objects.filter(id=id).first()
    if not user:
        return JsonResponse({"success": False, "message": "Пользователь не найден", "data": None})
    return JsonResponse({"success": True, "message": "Пользователь получен успешно", "data": UserSerializer(user).data})


@csrf_exempt
@require_http_methods(["POST"])
def users_create(request):
    payload = _json_body(request)
    roles = payload.get("roles") or []
    if not roles:
        return JsonResponse({"success": False, "message": "Необходимо указать хотя бы одну роль", "data": None})
    user = User.objects.create(
        username=(payload.get("username") or "").strip(),
        password=payload.get("password") or "",
        full_name=(payload.get("fullName") or "").strip(),
        email=(payload.get("email") or None),
        roles=roles,
    )
    return JsonResponse({"success": True, "message": "Пользователь создан успешно", "data": UserSerializer(user).data})


@csrf_exempt
@require_http_methods(["PUT"])
def users_update(request, id):
    user = get_object_or_404(User, id=id)
    payload = _json_body(request)
    user.username = (payload.get("username") or user.username).strip()
    if payload.get("password"):
        user.password = payload["password"]
    user.full_name = (payload.get("fullName") or user.full_name).strip()
    user.email = payload.get("email")
    user.roles = payload.get("roles") or user.roles
    user.save()
    return JsonResponse({"success": True, "message": "Пользователь обновлен успешно", "data": UserSerializer(user).data})


@csrf_exempt
@require_http_methods(["POST"])
def users_deactivate(request, id):
    user = get_object_or_404(User, id=id)
    payload = _json_body(request)
    user.enabled = False
    user.deactivation_reason = payload.get("reason") or "Деактивирован через API"
    user.save()
    return JsonResponse({"success": True, "message": "Пользователь деактивирован", "data": None})


@csrf_exempt
@require_http_methods(["POST"])
def users_activate(request, id):
    user = get_object_or_404(User, id=id)
    user.enabled = True
    user.deactivation_reason = None
    user.save()
    return JsonResponse({"success": True, "message": "Пользователь активирован", "data": None})


@csrf_exempt
@require_http_methods(["DELETE"])
def users_delete(request, id):
    User.objects.filter(id=id).delete()
    return JsonResponse({"success": True, "message": "Пользователь удален", "data": None})


@require_http_methods(["GET"])
def users_search(request):
    query = request.GET.get("query", "")
    users = User.objects.filter(
        Q(username__icontains=query) | Q(full_name__icontains=query) | Q(email__icontains=query)
    )
    return JsonResponse({"success": True, "message": "Поиск выполнен успешно", "data": UserSerializer(users, many=True).data})


@require_http_methods(["GET"])
def test_runs_by_project(request, project_id):
    test_runs = TestRun.objects.filter(project_id=project_id)
    return JsonResponse(TestRunSerializer(test_runs, many=True).data, safe=False)


@require_http_methods(["GET"])
def test_run_detailed(request, id):
    test_run = TestRun.objects.filter(id=id).first()
    if not test_run:
        return JsonResponse(build_api_response(False, "Тест-ран не найден"))
    return JsonResponse(build_api_response(True, testRun=TestRunSerializer(test_run).data))


@require_http_methods(["GET"])
def test_run_statistics(request, id):
    rows = TestRunTestCase.objects.filter(test_run_id=id)
    passed = rows.filter(status=TestRunTestCase.TestCaseStatus.PASSED).count()
    failed = rows.filter(status=TestRunTestCase.TestCaseStatus.FAILED).count()
    skipped = rows.filter(status=TestRunTestCase.TestCaseStatus.SKIPPED).count()
    not_run = rows.filter(status=TestRunTestCase.TestCaseStatus.NOT_RUN).count()
    return JsonResponse(
        build_api_response(
            True,
            passed=passed,
            failed=failed,
            skipped=skipped,
            not_run=not_run,
        )
    )


@csrf_exempt
@require_http_methods(["POST"])
def test_run_create(request, project_id):
    payload = _json_body(request)
    test_run = create_test_run(project_id, payload)
    return JsonResponse(build_api_response(True, "Тест-ран создан успешно!", testRun=TestRunSerializer(test_run).data))


@csrf_exempt
@require_http_methods(["PUT"])
def test_run_update(request, id):
    test_run = get_object_or_404(TestRun, id=id)
    payload = _json_body(request)
    test_run.title = payload.get("title") or test_run.title
    test_run.description = payload.get("description")
    test_run.executor_name = payload.get("executorName") or test_run.executor_name
    test_run.creator_name = payload.get("creatorName") or test_run.creator_name
    test_run.status = normalize_status(
        payload.get("status"),
        [v for v, _ in TestRun.Status.choices],
        test_run.status,
    )
    test_run.save()
    add_ids = payload.get("testCaseIdsToAdd") or []
    remove_ids = payload.get("testCaseIdsToRemove") or []
    for case in TestCase.objects.filter(id__in=add_ids):
        TestRunTestCase.objects.get_or_create(test_run=test_run, test_case=case)
    if remove_ids:
        TestRunTestCase.objects.filter(test_run=test_run, test_case_id__in=remove_ids).delete()
    return JsonResponse(build_api_response(True, "Тест-ран обновлен успешно!", testRun=TestRunSerializer(test_run).data))


@csrf_exempt
@require_http_methods(["PUT"])
def test_run_status_update(request, id):
    test_run = get_object_or_404(TestRun, id=id)
    payload = _json_body(request)
    test_run.status = normalize_status(
        payload.get("status"),
        [v for v, _ in TestRun.Status.choices],
        test_run.status,
    )
    test_run.save()
    return JsonResponse(build_api_response(True, testRun=TestRunSerializer(test_run).data))


@csrf_exempt
@require_http_methods(["PUT"])
def test_run_test_case_status_update(request, test_run_id, test_case_id):
    payload = _json_body(request)
    record = get_object_or_404(TestRunTestCase, test_run_id=test_run_id, test_case_id=test_case_id)
    core_user = _resolve_core_user(request)
    is_admin = _is_admin_user(core_user)

    record.status = normalize_status(
        payload.get("status"),
        [v for v, _ in TestRunTestCase.TestCaseStatus.choices],
        record.status,
    )
    incoming_comment = payload.get("comment")
    if incoming_comment is not None:
        new_comment = str(incoming_comment).strip() or None
        current_comment = (record.comment or "").strip() or None
        comment_is_changing = new_comment != current_comment
        has_legacy_comment_without_author = current_comment is not None and not record.comment_author_id

        if comment_is_changing and has_legacy_comment_without_author and not is_admin:
            return JsonResponse(
                build_api_response(
                    False,
                    "Комментарий без автора может изменить только администратор.",
                )
            )

        if comment_is_changing and record.comment_author_id and core_user and record.comment_author_id != core_user.id and not is_admin:
            return JsonResponse(
                build_api_response(
                    False,
                    "Редактировать комментарий может только его автор или администратор.",
                )
            )
        if comment_is_changing and record.comment_author_id and not core_user and not is_admin:
            return JsonResponse(
                build_api_response(
                    False,
                    "Для редактирования комментария требуется авторизация.",
                )
            )

        record.comment = new_comment
        if new_comment is None:
            record.comment_author = None
            record.comment_updated_at = None
        elif comment_is_changing:
            if not record.comment_author_id and current_comment is None:
                # New comment: first editor becomes owner.
                record.comment_author = core_user if core_user else record.comment_author
            elif not record.comment_author_id and current_comment is not None and is_admin:
                # Legacy comment without owner can be reassigned only by admin.
                record.comment_author = core_user if core_user else record.comment_author
            record.comment_updated_at = timezone.now()

    record.save()
    return JsonResponse(build_api_response(True, testRunTestCase=TestRunTestCaseSerializer(record).data))


@csrf_exempt
@require_http_methods(["DELETE"])
def test_run_delete(request, id):
    deleted, _ = TestRun.objects.filter(id=id).delete()
    return JsonResponse(build_api_response(bool(deleted), "Тест-ран удален успешно" if deleted else "Тест-ран не найден"))


@require_http_methods(["GET"])
def test_runs_search(request, project_id):
    term = request.GET.get("term", "")
    test_runs = TestRun.objects.filter(project_id=project_id, title__icontains=term)
    return JsonResponse(TestRunSerializer(test_runs, many=True).data, safe=False)


@csrf_exempt
@require_http_methods(["GET", "POST"])
def ai_analysis_run(request, test_suite_id):
    if request.method == "GET":
        latest = AIAnalysis.objects.filter(test_suite_id=test_suite_id).order_by("-created_at", "-id").first()
        if not latest:
            return JsonResponse(build_api_response(False, "Анализ для этого сьюта еще не запускался"), status=404)
        return JsonResponse(
            build_api_response(
                True,
                "Загружен последний анализ",
                analysis_id=latest.id,
                response=latest.ai_response,
                created_at=latest.created_at,
            )
        )

    core_user = _resolve_core_user(request)
    test_suite = TestSuite.objects.select_related("project").filter(id=test_suite_id).first()
    activity = _create_ai_activity_log(
        action_type=AIActivityLog.ActionType.ANALYZE_TEST_SUITE,
        status=AIActivityLog.Status.RUNNING,
        initiated_by=core_user,
        project=test_suite.project if test_suite else None,
        test_suite=test_suite,
        message="Запуск AI-анализа тест-сьюта",
    )
    result = run_ai_test_suite_analysis(test_suite_id)
    activity.status = AIActivityLog.Status.SUCCESS if result.get("success") else AIActivityLog.Status.FAILED
    activity.message = result.get("message") or activity.message
    activity.finished_at = timezone.now()
    activity.save(update_fields=["status", "message", "finished_at", "updated_at"])
    return JsonResponse(result)


@csrf_exempt
@require_http_methods(["POST"])
def ai_review_suite(request, test_suite_id):
    core_user = _resolve_core_user(request)
    if not core_user:
        return JsonResponse(build_api_response(False, "Нужна авторизация"), status=403)

    test_suite = TestSuite.objects.select_related("project").filter(id=test_suite_id).first()
    case_ids = list(TestCase.objects.filter(test_suite_id=test_suite_id).values_list("id", flat=True).order_by("id"))
    if not case_ids:
        _create_ai_activity_log(
            action_type=AIActivityLog.ActionType.REVIEW_TEST_SUITE,
            status=AIActivityLog.Status.FAILED,
            initiated_by=core_user,
            project=test_suite.project if test_suite else None,
            test_suite=test_suite,
            message="Запуск ревью отклонен: в тест-сьюте нет тест-кейсов",
        )
        return JsonResponse(build_api_response(False, "В тест-сьюте нет тест-кейсов"))

    with transaction.atomic():
        job, _ = TestSuiteAIReviewJob.objects.select_for_update().get_or_create(test_suite_id=test_suite_id)
        if job.status == TestSuiteAIReviewJob.Status.RUNNING:
            owner = job.started_by.full_name if job.started_by else "другой пользователь"
            _create_ai_activity_log(
                action_type=AIActivityLog.ActionType.REVIEW_TEST_SUITE,
                status=AIActivityLog.Status.FAILED,
                initiated_by=core_user,
                project=test_suite.project if test_suite else None,
                test_suite=test_suite,
                message=f"Запуск ревью отклонен: уже запущено ({owner})",
            )
            return JsonResponse(
                build_api_response(
                    False,
                    f"Ревью уже запущено ({owner}). Дождитесь завершения.",
                    job=_serialize_ai_review_job(job),
                ),
                status=409,
            )
        job.status = TestSuiteAIReviewJob.Status.RUNNING
        job.queue_case_ids = case_ids
        job.total_cases = len(case_ids)
        job.processed_cases = 0
        job.success_cases = 0
        job.failed_cases = 0
        job.started_by = core_user
        job.started_at = timezone.now()
        job.finished_at = None
        job.last_error = None
        job.save()
    _create_ai_activity_log(
        action_type=AIActivityLog.ActionType.REVIEW_TEST_SUITE,
        status=AIActivityLog.Status.RUNNING,
        initiated_by=core_user,
        project=test_suite.project if test_suite else None,
        test_suite=test_suite,
        message=f"Запущено массовое ревью ({len(case_ids)} тест-кейсов)",
    )
    return JsonResponse(build_api_response(True, "Очередь AI-ревью сформирована", job=_serialize_ai_review_job(job)))


@csrf_exempt
@require_http_methods(["POST"])
def ai_review_suite_queue_next(request, test_suite_id):
    core_user = _resolve_core_user(request)
    if not core_user:
        return JsonResponse(build_api_response(False, "Нужна авторизация"), status=403)

    with transaction.atomic():
        job = TestSuiteAIReviewJob.objects.select_for_update().filter(test_suite_id=test_suite_id).first()
        if not job or job.status != TestSuiteAIReviewJob.Status.RUNNING:
            return JsonResponse(build_api_response(False, "Очередь ревью не запущена"), status=409)

        if not job.queue_case_ids:
            job.status = TestSuiteAIReviewJob.Status.COMPLETED
            job.finished_at = timezone.now()
            job.save(update_fields=["status", "finished_at", "updated_at"])
            return JsonResponse(
                build_api_response(True, "Ревью завершено", completed=True, job=_serialize_ai_review_job(job))
            )

        case_id = int(job.queue_case_ids[0])
        job.queue_case_ids = job.queue_case_ids[1:]
        job.save(update_fields=["queue_case_ids", "updated_at"])

    # Run review outside lock/transaction to reduce lock time.
    review = create_or_update_review(case_id)

    with transaction.atomic():
        job = TestSuiteAIReviewJob.objects.select_for_update().get(test_suite_id=test_suite_id)
        job.processed_cases += 1
        if review:
            job.success_cases += 1
        else:
            job.failed_cases += 1
            job.last_error = f"Не удалось выполнить ревью для TC-{case_id}"

        completed = not job.queue_case_ids
        if completed:
            job.status = TestSuiteAIReviewJob.Status.COMPLETED
            job.finished_at = timezone.now()
        job.save()

    if completed:
        suite_activity = (
            AIActivityLog.objects.filter(
                action_type=AIActivityLog.ActionType.REVIEW_TEST_SUITE,
                test_suite_id=test_suite_id,
                status=AIActivityLog.Status.RUNNING,
                finished_at__isnull=True,
            )
            .order_by("-started_at", "-id")
            .first()
        )
        if suite_activity:
            suite_activity.status = (
                AIActivityLog.Status.SUCCESS if job.failed_cases == 0 else AIActivityLog.Status.FAILED
            )
            suite_activity.message = (
                f"Массовое ревью завершено: обработано {job.processed_cases}/{job.total_cases}, "
                f"успешно {job.success_cases}, ошибки {job.failed_cases}"
            )
            suite_activity.finished_at = timezone.now()
            suite_activity.save(update_fields=["status", "message", "finished_at", "updated_at"])

    return JsonResponse(
        build_api_response(
            True,
            "Шаг очереди выполнен",
            completed=completed,
            review=TestCaseReviewSerializer(review).data if review else None,
            job=_serialize_ai_review_job(job),
        )
    )


@require_http_methods(["GET"])
def ai_review_suite_queue_status(request, test_suite_id):
    job = TestSuiteAIReviewJob.objects.filter(test_suite_id=test_suite_id).first()
    if not job:
        # No active queue is an expected idle state; keep HTTP 200
        # to avoid filling logs with repeated "Not Found" entries.
        return JsonResponse(build_api_response(False, "Очередь не запущена", job=None))
    return JsonResponse(build_api_response(True, job=_serialize_ai_review_job(job)))


@csrf_exempt
@require_http_methods(["POST"])
def ai_review_case(request, test_case_id):
    core_user = _resolve_core_user(request)
    test_case = TestCase.objects.select_related("test_suite__project").filter(id=test_case_id).first()
    activity = _create_ai_activity_log(
        action_type=AIActivityLog.ActionType.REVIEW_TEST_CASE,
        status=AIActivityLog.Status.RUNNING,
        initiated_by=core_user,
        project=test_case.test_suite.project if test_case and test_case.test_suite else None,
        test_suite=test_case.test_suite if test_case else None,
        test_case=test_case,
        message="Запуск AI ревью тест-кейса",
    )
    review = create_or_update_review(test_case_id)
    activity.status = AIActivityLog.Status.SUCCESS if review else AIActivityLog.Status.FAILED
    activity.message = (
        "AI ревью завершено" if review else "AI ревью не выполнено"
    )
    activity.finished_at = timezone.now()
    activity.save(update_fields=["status", "message", "finished_at", "updated_at"])
    if not review:
        return JsonResponse(build_api_response(False, "Тест-кейс не найден"))
    return JsonResponse(build_api_response(True, "AI ревью завершено", review=TestCaseReviewSerializer(review).data))


@require_http_methods(["GET"])
def ai_review_case_get(request, test_case_id):
    review = TestCaseReview.objects.filter(test_case_id=test_case_id).first()
    if not review:
        return JsonResponse(build_api_response(False, "Ревью не найдено"))
    return JsonResponse(build_api_response(True, review=TestCaseReviewSerializer(review).data))


@require_http_methods(["GET"])
def ai_review_suite_get(request, test_suite_id):
    reviews = TestCaseReview.objects.filter(test_case__test_suite_id=test_suite_id)
    return JsonResponse(build_api_response(True, reviews=TestCaseReviewSerializer(reviews, many=True).data))


@csrf_exempt
@require_http_methods(["DELETE"])
def ai_review_case_delete(request, test_case_id):
    deleted, _ = TestCaseReview.objects.filter(test_case_id=test_case_id).delete()
    return JsonResponse(build_api_response(bool(deleted), "Ревью удалено" if deleted else "Ревью не найдено"))


@require_http_methods(["GET"])
def token_status(request):
    return JsonResponse(
        {
            "tokenValid": True,
            "expiryTime": None,
            "currentTime": None,
        }
    )


@require_http_methods(["GET"])
def token_refresh(request):
    return JsonResponse({"status": "Token refresh initiated"})


@csrf_exempt
@require_http_methods(["POST"])
def admin_ai_test_connection(request):
    core_user = _resolve_core_user(request)
    if not _is_admin_user(core_user):
        return JsonResponse(build_api_response(False, "Недостаточно прав для проверки подключения"), status=403)
    result = test_ai_provider_connection()
    return JsonResponse(result, status=200 if result.get("success") else 400)


@csrf_exempt
@require_http_methods(["GET", "POST"])
def traceability_ai_quality(request, project_id):
    force_refresh = request.method == "POST"
    project = Project.objects.filter(id=project_id).first()
    core_user = _resolve_core_user(request)
    activity = None
    if force_refresh:
        activity = _create_ai_activity_log(
            action_type=AIActivityLog.ActionType.ANALYZE_TRACEABILITY_MODEL,
            status=AIActivityLog.Status.RUNNING,
            initiated_by=core_user,
            project=project,
            message=f"Запуск AI-оценки тестовой модели проекта #{project_id}",
        )
    result = analyze_traceability_model_quality(project_id, force_refresh=force_refresh)
    if activity:
        activity.status = AIActivityLog.Status.SUCCESS if result.get("success") else AIActivityLog.Status.FAILED
        activity.message = result.get("message") or (
            "AI-оценка тестовой модели сформирована" if result.get("success") else "AI-оценка тестовой модели завершилась с ошибкой"
        )
        activity.finished_at = timezone.now()
        activity.save(update_fields=["status", "message", "finished_at", "updated_at"])
    return JsonResponse(result, status=200 if result.get("success") else 400)


@require_http_methods(["GET"])
def traceability_matrix_export_excel(request, project_id):
    project = Project.objects.filter(id=project_id).first()
    if not project:
        return JsonResponse(build_api_response(False, "Проект не найден"), status=404)

    user_stories = UserStory.objects.filter(section__project_id=project_id).order_by("id")
    test_cases = TestCase.objects.filter(test_suite__project_id=project_id).order_by("id")
    links = set(
        TestCaseUserStory.objects.filter(
            test_case_id__in=test_cases.values_list("id", flat=True),
            user_story_id__in=user_stories.values_list("id", flat=True),
        ).values_list("test_case_id", "user_story_id")
    )

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError:
        csv_buffer = StringIO()
        writer = csv.writer(csv_buffer, delimiter=";")
        writer.writerow(["User Story \\ Test Case", *[f"TC-{tc.id}" for tc in test_cases]])
        for us in user_stories:
            row = [f"US-{us.id}: {us.name}"]
            for tc in test_cases:
                row.append("✓" if (tc.id, us.id) in links else "")
            writer.writerow(row)
        response = HttpResponse(csv_buffer.getvalue(), content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = (
            f'attachment; filename="coverage-matrix-project-{project_id}.csv"'
        )
        return response

    wb = Workbook()
    ws = wb.active
    ws.title = "Coverage Matrix"

    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.cell(row=1, column=1, value="User Story \\ Test Case")
    for col_idx, tc in enumerate(test_cases, start=2):
        ws.cell(row=1, column=col_idx, value=f"TC-{tc.id}")
    for col_idx in range(1, len(test_cases) + 2):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment

    for row_idx, us in enumerate(user_stories, start=2):
        ws.cell(row=row_idx, column=1, value=f"US-{us.id}: {us.name}")
        ws.cell(row=row_idx, column=1).alignment = Alignment(vertical="top", wrap_text=True)
        for col_idx, tc in enumerate(test_cases, start=2):
            ws.cell(row=row_idx, column=col_idx, value="✓" if (tc.id, us.id) in links else "")
            ws.cell(row=row_idx, column=col_idx).alignment = center_alignment

    ws.freeze_panes = "B2"
    ws.column_dimensions["A"].width = 56
    for col_idx in range(2, len(test_cases) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 10

    filename = f"coverage-matrix-project-{project_id}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    wb.close()
    return response


@csrf_exempt
def projects_collection(request):
    if request.method == "GET":
        return projects_list(request)
    if request.method == "POST":
        return projects_create(request)
    return JsonResponse({}, status=405)


@csrf_exempt
def projects_item(request, id):
    if request.method == "GET":
        return projects_get(request, id)
    if request.method == "PUT":
        return projects_update(request, id)
    if request.method == "DELETE":
        return projects_delete(request, id)
    return JsonResponse({}, status=405)


@csrf_exempt
def project_test_suites_collection(request, project_id):
    if request.method == "GET":
        return project_test_suites(request, project_id)
    if request.method == "POST":
        return project_test_suites_create(request, project_id)
    return JsonResponse({}, status=405)


@csrf_exempt
def project_sections_collection(request, project_id):
    if request.method == "GET":
        return project_sections(request, project_id)
    if request.method == "POST":
        return project_sections_create(request, project_id)
    return JsonResponse({}, status=405)


@csrf_exempt
def project_tags_collection(request, project_id):
    if request.method == "GET":
        return project_tags(request, project_id)
    if request.method == "POST":
        return project_tags_create(request, project_id)
    return JsonResponse({}, status=405)


@csrf_exempt
def project_test_runs_collection(request, project_id):
    if request.method == "GET":
        return test_runs_by_project(request, project_id)
    if request.method == "POST":
        return test_run_create(request, project_id)
    return JsonResponse({}, status=405)


@csrf_exempt
def test_suite_item(request, id):
    if request.method == "GET":
        return test_suite_get(request, id)
    if request.method == "PUT":
        return test_suite_update(request, id)
    if request.method == "DELETE":
        return test_suite_delete(request, id)
    return JsonResponse({}, status=405)


@csrf_exempt
def test_suite_cases_collection(request, test_suite_id):
    if request.method == "GET":
        return test_suite_cases(request, test_suite_id)
    if request.method == "POST":
        return test_suite_cases_create(request, test_suite_id)
    return JsonResponse({}, status=405)


@csrf_exempt
def test_case_item(request, id):
    if request.method == "GET":
        return test_case_get(request, id)
    if request.method == "PUT":
        return test_case_update(request, id)
    if request.method == "DELETE":
        return test_case_delete(request, id)
    return JsonResponse({}, status=405)


@csrf_exempt
def test_case_tag_item(request, test_case_id, tag_id):
    if request.method == "POST":
        return test_case_add_tag(request, test_case_id, tag_id)
    if request.method == "DELETE":
        return test_case_remove_tag(request, test_case_id, tag_id)
    return JsonResponse({}, status=405)


@csrf_exempt
def test_step_item(request, step_id):
    if request.method == "PUT":
        return test_step_update(request, step_id)
    if request.method == "DELETE":
        return test_step_delete(request, step_id)
    return JsonResponse({}, status=405)


@csrf_exempt
def test_case_user_story_item(request, test_case_id, user_story_id):
    if request.method == "POST":
        return test_case_add_user_story(request, test_case_id, user_story_id)
    if request.method == "DELETE":
        return test_case_remove_user_story(request, test_case_id, user_story_id)
    return JsonResponse({}, status=405)


@csrf_exempt
def section_item(request, id):
    if request.method == "GET":
        return section_get(request, id)
    if request.method == "PUT":
        return section_update(request, id)
    if request.method == "DELETE":
        return section_delete(request, id)
    return JsonResponse({}, status=405)


@csrf_exempt
def section_user_stories_collection(request, section_id):
    if request.method == "GET":
        return section_user_stories(request, section_id)
    if request.method == "POST":
        return section_user_stories_create(request, section_id)
    return JsonResponse({}, status=405)


@csrf_exempt
def user_story_item(request, id):
    if request.method == "GET":
        return user_story_get(request, id)
    if request.method == "PUT":
        return user_story_update(request, id)
    if request.method == "DELETE":
        return user_story_delete(request, id)
    return JsonResponse({}, status=405)


@csrf_exempt
def tag_item(request, id):
    if request.method == "GET":
        return tag_get(request, id)
    if request.method == "PUT":
        return tag_update(request, id)
    if request.method == "DELETE":
        return tag_delete(request, id)
    return JsonResponse({}, status=405)


@csrf_exempt
def comment_item(request, comment_id):
    if request.method == "PUT":
        return comments_update(request, comment_id)
    if request.method == "DELETE":
        return comments_delete(request, comment_id)
    return JsonResponse({}, status=405)


@csrf_exempt
def users_collection(request):
    if request.method == "GET":
        return users_list(request)
    if request.method == "POST":
        return users_create(request)
    return JsonResponse({}, status=405)


@csrf_exempt
def user_item(request, id):
    if request.method == "GET":
        return users_get(request, id)
    if request.method == "PUT":
        return users_update(request, id)
    if request.method == "DELETE":
        return users_delete(request, id)
    return JsonResponse({}, status=405)


@csrf_exempt
def test_run_item(request, id):
    if request.method == "PUT":
        return test_run_update(request, id)
    if request.method == "DELETE":
        return test_run_delete(request, id)
    return JsonResponse({}, status=405)


@csrf_exempt
def test_case_ai_review_item(request, test_case_id):
    if request.method == "POST":
        return ai_review_case(request, test_case_id)
    if request.method == "GET":
        return ai_review_case_get(request, test_case_id)
    if request.method == "DELETE":
        return ai_review_case_delete(request, test_case_id)
    return JsonResponse({}, status=405)
